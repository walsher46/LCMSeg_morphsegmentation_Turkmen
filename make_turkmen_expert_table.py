from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INSTRUCTION = (
    "Инструкция: заполните колонку D (TRUE / FALSE / PARTIAL). "
    "Если FALSE или PARTIAL — напишите правильную сегментацию в колонке E "
    "через @@ (например: adam@@ lar@@ yň)"
)

HEADERS = [
    "Слово",
    "Сегментация FEMSeg",
    "Морфем",
    'Верно?\n(TRUE/FALSE/PARTIAL)',
    "Правильная сегментация",
    "Комментарий",
]

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def normalize_segmented_line(line: str) -> str:
    line = line.strip()
    line = re.sub(r"\s+", " ", line)
    return line


def extract_segmented_words(line: str) -> List[str]:
    """
    Разбивает строку на токены, сохраняя формы с @@ как один токен.
    Пунктуацию не включает в таблицу.
    Примеры:
      'aralygy@@ nda' -> один токен
      'merý' -> один токен
    """
    tokens = line.split()
    words = []
    i = 0
    punct_only = {",", ".", ":", ";", "!", "?", "%", '"', "'", "(", ")", "[", "]", "{", "}"}

    while i < len(tokens):
        tok = tokens[i]

        if tok in punct_only or re.fullmatch(r"[-–—]+", tok):
            i += 1
            continue

        current = [tok]
        while current[-1].endswith("@@") and i + 1 < len(tokens):
            i += 1
            nxt = tokens[i]
            current.append(nxt)
            if nxt in punct_only:
                break

        joined = " ".join(current).strip()
        joined = re.sub(r"\s+([,.:;!?%])", r"\1", joined)

        clean_for_check = joined.replace("@@", "").strip("-–—")
        if clean_for_check and not re.fullmatch(r"[\W_]+", clean_for_check, flags=re.UNICODE):
            words.append(joined)
        i += 1

    return words


def morph_count(segmented_word: str) -> int:
    return segmented_word.count("@@") + 1 if segmented_word else 0


def plain_word(segmented_word: str) -> str:
    return segmented_word.replace("@@", "")


def iter_input_lines(path: Path) -> Iterable[str]:
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = normalize_segmented_line(line)
            if line:
                yield line


def build_workbook(input_path: Path, output_path: Path, sheet_name: str = "Expert Review") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.freeze_panes = "A3"

    # Строка инструкции
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = INSTRUCTION
    c.font = Font(name="Calibri", size=13, bold=False, italic=True, color="6B5B00")
    c.fill = PatternFill("solid", fgColor="FFF2CC")
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = BORDER
    ws.row_dimensions[1].height = 34

    # Заголовки
    header_fill = PatternFill("solid", fgColor="1F3B5C")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    ws.row_dimensions[2].height = 32

    row = 3
    for line in iter_input_lines(input_path):
        segmented_words = extract_segmented_words(line)
        for seg in segmented_words:
            ws.cell(row=row, column=1, value=plain_word(seg))
            ws.cell(row=row, column=2, value=seg)
            ws.cell(row=row, column=3, value=morph_count(seg))
            ws.cell(row=row, column=4, value="")
            ws.cell(row=row, column=5, value="")
            ws.cell(row=row, column=6, value="")
            row += 1

    # Стили данных
    for r in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in r:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Ширины колонок
    widths = {
        "A": 28,
        "B": 36,
        "C": 12,
        "D": 18,
        "E": 38,
        "F": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Фильтр
    ws.auto_filter.ref = f"A2:F{ws.max_row}"

    # Выпадающий список для D
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"TRUE,FALSE,PARTIAL"', allow_blank=True)
    dv.promptTitle = "Оценка сегментации"
    dv.prompt = "Выберите TRUE, FALSE или PARTIAL"
    ws.add_data_validation(dv)
    dv.add(f"D3:D{ws.max_row}")

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Создает Excel-таблицу для экспертной оценки сегментации туркменского текста."
    )
    parser.add_argument(
        "-i", "--input",
        required=True,
        help="Путь к входному TXT-файлу с сегментированным текстом",
    )
    parser.add_argument(
        "-o", "--output",
        default="turkmen_expert_table.xlsx",
        help="Путь к выходному XLSX-файлу",
    )
    parser.add_argument(
        "--sheet-name",
        default="Expert Review",
        help="Имя листа Excel",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    build_workbook(
        input_path=Path(args.input),
        output_path=Path(args.output),
        sheet_name=args.sheet_name,
    )
    print(f"Готово: {args.output}")
